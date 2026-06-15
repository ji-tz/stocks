"""回测结果导出模块：提供 Excel 和 PDF 数据导出功能。

职责：
 - 将回测结果（BacktestResult dict）导出为 .xlsx 文件
 - 为 PDF 生成准备结构化数据

由 TRADER 维护。
"""

import os
from typing import Any, Dict, List, Optional


def export_to_excel(
    backtest_result: Dict[str, Any],
    output_path: str,
    strategy_name: str = "",
) -> str:
    """将回测结果导出为 .xlsx 文件。

    Args:
        backtest_result: simulator.run/simulate 返回的完整回测结果字典。
        output_path: 保存路径（包含 .xlsx 扩展名）。
        strategy_name: 策略名称，用于文件命名和标识。

    Returns:
        保存的文件路径。

    Raises:
        ImportError: 如果 openpyxl 未安装。
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    symbol = backtest_result.get("symbol", "UNKNOWN")
    start_date = backtest_result.get("start_date", "")
    end_date = backtest_result.get("end_date", "")
    init_cash = backtest_result.get("init_cash", 0)
    final_cash = backtest_result.get("cash", 0)
    total_value = backtest_result.get("total_value", 0)
    realized_pl = backtest_result.get("realized_pl", 0)
    unrealized_pl = backtest_result.get("unrealized_pl", 0)
    shares = backtest_result.get("shares", 0)
    last_price = backtest_result.get("last_price", 0)
    trades_count = backtest_result.get("trades", 0)

    # 样式定义
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def style_header(ws, cols: int):
        for col in range(1, cols + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

    def auto_width(ws, cols: int, max_width: int = 30):
        for col in range(1, cols + 1):
            letter = get_column_letter(col)
            max_len = 0
            for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
                for cell in row:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[letter].width = min(max_len + 3, max_width)

    # ========== Sheet 1: 交易明细 ==========
    ws_trades = wb.active
    ws_trades.title = "交易明细"

    trade_headers = ["日期", "方向", "成交价", "数量(股)", "成交后现金", "成交后持仓", "已实现盈亏", "备注"]
    for col, h in enumerate(trade_headers, 1):
        ws_trades.cell(row=1, column=col, value=h)

    trades_list = backtest_result.get("trades_list", [])
    for i, t in enumerate(trades_list, 2):
        ws_trades.cell(row=i, column=1, value=t.get("date", ""))
        ws_trades.cell(row=i, column=2, value="买入" if t.get("action") == "buy" else "卖出")
        ws_trades.cell(row=i, column=3, value=t.get("price", 0))
        ws_trades.cell(row=i, column=4, value=t.get("shares", 0))
        ws_trades.cell(row=i, column=5, value=t.get("cash", 0))
        ws_trades.cell(row=i, column=6, value=t.get("shares_after", 0))
        ws_trades.cell(row=i, column=7, value=t.get("realized_pl", ""))
        ws_trades.cell(row=i, column=8, value=t.get("source", ""))

    style_header(ws_trades, len(trade_headers))
    auto_width(ws_trades, len(trade_headers))

    # ========== Sheet 2: 持仓变化 ==========
    ws_history = wb.create_sheet("持仓变化")

    history_headers = ["日期", "现金", "持仓(股)", "平均成本", "最新价", "市值", "总资产"]
    for col, h in enumerate(history_headers, 1):
        ws_history.cell(row=1, column=col, value=h)

    history = backtest_result.get("history", [])
    for i, h in enumerate(history, 2):
        ws_history.cell(row=i, column=1, value=h.get("date", ""))
        ws_history.cell(row=i, column=2, value=h.get("cash", 0))
        ws_history.cell(row=i, column=3, value=h.get("shares", 0))
        ws_history.cell(row=i, column=4, value=h.get("avg_cost", 0))
        ws_history.cell(row=i, column=5, value=h.get("last_price", 0))
        ws_history.cell(row=i, column=6, value=h.get("market_value", 0))
        ws_history.cell(row=i, column=7, value=h.get("total_value", 0))

    style_header(ws_history, len(history_headers))
    auto_width(ws_history, len(history_headers))

    # ========== Sheet 3: 汇总统计 ==========
    ws_summary = wb.create_sheet("汇总统计")

    total_pl = realized_pl + unrealized_pl
    total_return_rate = (total_pl / init_cash * 100) if init_cash else 0

    summary_data = [
        ("股票代码", symbol),
        ("策略名称", strategy_name or "N/A"),
        ("开始日期", start_date),
        ("结束日期", end_date),
        ("初始资金", init_cash),
        ("最终现金", round(final_cash, 2)),
        ("期末持仓(股)", shares),
        ("最新价格", last_price),
        ("持仓市值", round(total_value - final_cash, 2) if total_value >= final_cash else 0),
        ("总资产", round(total_value, 2)),
        ("已实现盈亏", round(realized_pl, 2)),
        ("未实现盈亏", round(unrealized_pl, 2)),
        ("总盈亏", round(total_pl, 2)),
        ("总收益率(%)", round(total_return_rate, 2)),
        ("交易次数", trades_count),
    ]

    key_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    for i, (key, val) in enumerate(summary_data, 1):
        cell_k = ws_summary.cell(row=i, column=1, value=key)
        cell_v = ws_summary.cell(row=i, column=2, value=val)
        cell_k.font = header_font
        cell_k.fill = key_fill
        cell_k.border = thin_border
        cell_v.border = thin_border

    ws_summary.column_dimensions["A"].width = 18
    ws_summary.column_dimensions["B"].width = 20

    # 保存
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    return output_path


def prepare_pdf_data(
    backtest_result: Dict[str, Any],
    strategy_name: str = "",
) -> Dict[str, Any]:
    """从回测结果提取 PDF 所需的结构化数据。

    Args:
        backtest_result: simulator.run/simulate 返回的完整回测结果字典。
        strategy_name: 策略名称。

    Returns:
        包含以下键的字典：
        - summary: 概要信息（symbol, strategy_name, start_date, end_date, init_cash）
        - metrics: 关键指标（total_return_rate, annualized_return, max_drawdown, sharpe_ratio, total_pl）
        - trades: 交易清单
    """
    symbol = backtest_result.get("symbol", "UNKNOWN")
    start_date = backtest_result.get("start_date", "")
    end_date = backtest_result.get("end_date", "")
    init_cash = backtest_result.get("init_cash", 0)
    final_cash = backtest_result.get("cash", 0)
    total_value = backtest_result.get("total_value", 0)
    realized_pl = backtest_result.get("realized_pl", 0)
    unrealized_pl = backtest_result.get("unrealized_pl", 0)
    trades_list = backtest_result.get("trades_list", [])
    history = backtest_result.get("history", [])

    total_pl = realized_pl + unrealized_pl
    total_return_rate = (total_pl / init_cash) if init_cash else 0

    # 计算年化收益率
    annualized_return = 0.0
    if start_date and end_date and init_cash > 0:
        try:
            from datetime import datetime
            s = datetime.strptime(start_date, "%Y-%m-%d")
            e = datetime.strptime(end_date, "%Y-%m-%d")
            days = (e - s).days
            if days > 0:
                annualized_return = (
                    (total_value / init_cash) ** (365.0 / days) - 1
                )
        except (ValueError, ZeroDivisionError):
            annualized_return = 0.0

    # 计算最大回撤（从 history 计算）
    max_drawdown = 0.0
    if history:
        peak = history[0].get("total_value", init_cash) if history else init_cash
        peak = max(peak, init_cash)
        for h in history:
            tv = h.get("total_value", init_cash)
            if tv > peak:
                peak = tv
            drawdown = (peak - tv) / peak if peak > 0 else 0
            if drawdown > max_drawdown:
                max_drawdown = drawdown

    # 近似夏普比率（简单估算）
    sharpe_ratio = 0.0
    if history and len(history) >= 2:
        try:
            daily_returns = []
            for i in range(1, len(history)):
                prev_tv = history[i - 1].get("total_value", init_cash)
                curr_tv = history[i].get("total_value", init_cash)
                if prev_tv > 0:
                    daily_returns.append((curr_tv - prev_tv) / prev_tv)
            if daily_returns:
                import statistics
                avg_return = statistics.mean(daily_returns)
                std_return = statistics.stdev(daily_returns) if len(daily_returns) > 1 else 0
                if std_return > 0:
                    # 假设无风险利率 0.02
                    sharpe_ratio = (avg_return * 252 - 0.02) / (std_return * (252 ** 0.5))
        except (ValueError, ZeroDivisionError, IndexError):
            sharpe_ratio = 0.0

    return {
        "summary": {
            "symbol": symbol,
            "strategy_name": strategy_name,
            "start_date": start_date,
            "end_date": end_date,
            "init_cash": init_cash,
        },
        "metrics": {
            "total_return_rate": round(total_return_rate, 4),
            "annualized_return": round(annualized_return, 4),
            "max_drawdown": round(max_drawdown, 4),
            "sharpe_ratio": round(sharpe_ratio, 4),
            "total_pl": round(total_pl, 2),
            "final_value": round(total_value, 2),
        },
        "trades": [
            {
                "date": t.get("date", ""),
                "action": "买入" if t.get("action") == "buy" else "卖出",
                "price": t.get("price", 0),
                "shares": t.get("shares", 0),
                "pl": t.get("realized_pl", ""),
            }
            for t in trades_list
        ],
    }


def generate_filename(
    symbol: str,
    strategy_name: str,
    start_date: str,
    end_date: str,
    ext: str = "xlsx",
) -> str:
    """生成导出文件名。

    Args:
        symbol: 股票代码。
        strategy_name: 策略名称。
        start_date: 开始日期（YYYY-MM-DD）。
        end_date: 结束日期（YYYY-MM-DD）。
        ext: 文件扩展名（默认 xlsx）。

    Returns:
        格式化文件名。
    """
    # 清理策略名称中的非法字符
    safe_name = strategy_name.replace("/", "_").replace("\\", "_") if strategy_name else "backtest"
    date_part = f"{start_date}_{end_date}" if start_date and end_date else ""
    parts = [s for s in [symbol, safe_name, date_part] if s]
    return f"{'_'.join(parts)}.{ext.lstrip('.')}"
