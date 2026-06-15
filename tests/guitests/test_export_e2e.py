"""GUI 端到端测试：回测结果导出功能（Excel/PDF）。

验证 Flask 的 /download/excel 和 /download/pdf 路由能正确返回导出文件。

Owned by: GTEST (per AGENTS.md)
"""

import json
import unittest
from unittest.mock import patch


class TestExportRoutes(unittest.TestCase):
    """Test /download/excel and /download/pdf routes via Flask test client."""

    def setUp(self):
        from gui.web import app

        # Save original testing flag
        self._original_testing = app.testing
        app.testing = True
        self.client = app.test_client()

    def tearDown(self):
        from gui.web import app

        app.testing = self._original_testing

    def _make_result(self) -> dict:
        return {
            "symbol": "600900",
            "start_date": "2023-01-03",
            "end_date": "2023-01-31",
            "init_cash": 100000.0,
            "cash": 75900.0,
            "total_value": 84700.0,
            "realized_pl": 200.0,
            "unrealized_pl": 8800.0,
            "shares": 400,
            "last_price": 22.0,
            "trades": 3,
            "trades_list": [
                {"date": "2023-01-03", "action": "buy", "price": 20.0, "shares": 300,
                 "cash": 94000.0, "shares_after": 300, "realized_pl": 0, "source": "signal"},
                {"date": "2023-01-10", "action": "buy", "price": 21.0, "shares": 200,
                 "cash": 73900.0, "shares_after": 500, "realized_pl": 0, "source": "signal"},
                {"date": "2023-01-15", "action": "sell", "price": 22.0, "shares": 100,
                 "cash": 75900.0, "shares_after": 400, "realized_pl": 200.0, "source": "signal"},
            ],
            "history": [
                {"date": "2023-01-03", "cash": 94000.0, "shares": 300, "avg_cost": 20.0,
                 "last_price": 20.5, "market_value": 6150.0, "total_value": 100150.0},
                {"date": "2023-01-10", "cash": 73900.0, "shares": 500, "avg_cost": 20.4,
                 "last_price": 21.0, "market_value": 10500.0, "total_value": 84400.0},
                {"date": "2023-01-15", "cash": 75900.0, "shares": 400, "avg_cost": 20.5,
                 "last_price": 22.0, "market_value": 8800.0, "total_value": 84700.0},
            ],
        }

    def test_download_excel_returns_xlsx(self):
        """Verify /download/excel returns a valid Excel file."""
        result = self._make_result()
        rv = self.client.post(
            "/download/excel",
            data={
                "result_json": json.dumps(result),
                "strategy_name": "mean_cost",
            },
        )
        self.assertEqual(rv.status_code, 200)
        # Check content type
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            rv.content_type,
        )
        # Check Content-Disposition header for filename
        disposition = rv.headers.get("Content-Disposition", "")
        self.assertIn("attachment", disposition)
        self.assertIn(".xlsx", disposition)
        # Verify it's a valid .xlsx file
        self.assertGreater(len(rv.data), 500)

    def test_download_excel_opens_as_valid_xlsx(self):
        """Verify /download/excel response can be parsed as openpyxl workbook."""
        result = self._make_result()
        rv = self.client.post(
            "/download/excel",
            data={
                "result_json": json.dumps(result),
                "strategy_name": "sma",
            },
        )
        self.assertEqual(rv.status_code, 200)

        import openpyxl
        from io import BytesIO

        wb = openpyxl.load_workbook(BytesIO(rv.data))
        self.assertEqual(wb.sheetnames, ["交易明细", "持仓变化", "汇总统计"])
        wb.close()

    def test_download_excel_contains_trade_data(self):
        """Verify /download/excel response contains expected trade data."""
        result = self._make_result()
        rv = self.client.post(
            "/download/excel",
            data={
                "result_json": json.dumps(result),
                "strategy_name": "mean_cost",
            },
        )
        self.assertEqual(rv.status_code, 200)

        import openpyxl
        from io import BytesIO

        wb = openpyxl.load_workbook(BytesIO(rv.data))
        ws = wb["交易明细"]
        # Header row
        self.assertEqual(ws.cell(1, 1).value, "日期")
        self.assertEqual(ws.cell(1, 2).value, "方向")
        # Some trade data should exist
        self.assertGreater(ws.max_row, 1)
        wb.close()

    def test_download_excel_without_strategy_name(self):
        """Verify /download/excel works without strategy_name (uses default)."""
        result = self._make_result()
        rv = self.client.post(
            "/download/excel",
            data={"result_json": json.dumps(result)},
        )
        self.assertEqual(rv.status_code, 200)
        self.assertGreater(len(rv.data), 500)

    def test_download_excel_missing_result(self):
        """Verify /download/excel returns 400 when result_json is missing."""
        rv = self.client.post("/download/excel", data={})
        self.assertEqual(rv.status_code, 400)
        data = rv.get_json()
        self.assertIn("error", data)

    def test_download_excel_invalid_json(self):
        """Verify /download/excel returns 400 for invalid JSON."""
        rv = self.client.post(
            "/download/excel",
            data={"result_json": "not valid json"},
        )
        self.assertEqual(rv.status_code, 400)
        data = rv.get_json()
        self.assertIn("error", data)

    def test_download_excel_non_dict_json(self):
        """Verify /download/excel returns 400 when JSON is not a dict."""
        rv = self.client.post(
            "/download/excel",
            data={"result_json": json.dumps([1, 2, 3])},
        )
        self.assertEqual(rv.status_code, 400)
        data = rv.get_json()
        self.assertIn("error", data)

    def test_download_pdf_returns_pdf(self):
        """Verify /download/pdf returns a valid PDF file."""
        result = self._make_result()
        rv = self.client.post(
            "/download/pdf",
            data={
                "result_json": json.dumps(result),
                "strategy_name": "mean_cost",
            },
        )
        self.assertEqual(rv.status_code, 200)
        # PDF content type
        self.assertIn("application/pdf", rv.content_type)
        # Check Content-Disposition
        disposition = rv.headers.get("Content-Disposition", "")
        self.assertIn("attachment", disposition)
        self.assertIn(".pdf", disposition)
        # Verify it's a valid PDF (starts with %PDF)
        self.assertTrue(rv.data.startswith(b"%PDF"))
        self.assertGreater(len(rv.data), 100)

    def test_download_pdf_contains_report_content(self):
        """Verify /download/pdf response contains meaningful report content."""
        result = self._make_result()
        rv = self.client.post(
            "/download/pdf",
            data={
                "result_json": json.dumps(result),
                "strategy_name": "mean_cost",
            },
        )
        self.assertEqual(rv.status_code, 200)

        # Verify it's a valid PDF
        self.assertTrue(rv.data.startswith(b"%PDF"))
        # PDF should be at least 1KB with real content
        self.assertGreater(len(rv.data), 1000)

        # Check Content-Disposition header for filename parts
        disposition = rv.headers.get("Content-Disposition", "")
        self.assertIn("600900", disposition)
        self.assertIn("mean_cost", disposition)

    def test_download_pdf_without_strategy_name(self):
        """Verify /download/pdf works without strategy_name."""
        result = self._make_result()
        rv = self.client.post(
            "/download/pdf",
            data={"result_json": json.dumps(result)},
        )
        self.assertEqual(rv.status_code, 200)
        self.assertTrue(rv.data.startswith(b"%PDF"))

    def test_download_pdf_missing_result(self):
        """Verify /download/pdf returns 400 without result_json."""
        rv = self.client.post("/download/pdf", data={})
        self.assertEqual(rv.status_code, 400)
        data = rv.get_json()
        self.assertIn("error", data)

    def test_download_pdf_invalid_json(self):
        """Verify /download/pdf returns 400 for invalid JSON."""
        rv = self.client.post(
            "/download/pdf",
            data={"result_json": "bad json"},
        )
        self.assertEqual(rv.status_code, 400)
        data = rv.get_json()
        self.assertIn("error", data)

    def test_download_pdf_non_dict_json(self):
        """Verify /download/pdf returns 400 when JSON is not a dict."""
        rv = self.client.post(
            "/download/pdf",
            data={"result_json": json.dumps("string_value")},
        )
        self.assertEqual(rv.status_code, 400)
        data = rv.get_json()
        self.assertIn("error", data)

    def test_download_pdf_with_sample_data(self):
        """Verify /download/pdf handles a realistic but minimal result."""
        minimal_result = {"symbol": "600519", "total_value": 200000.0, "cash": 200000.0}
        rv = self.client.post(
            "/download/pdf",
            data={
                "result_json": json.dumps(minimal_result),
                "strategy_name": "sma",
            },
        )
        self.assertEqual(rv.status_code, 200)
        self.assertTrue(rv.data.startswith(b"%PDF"))

    def test_download_excel_with_sample_data(self):
        """Verify /download/excel handles a realistic but minimal result."""
        minimal_result = {"symbol": "600519", "total_value": 200000.0, "cash": 200000.0}
        rv = self.client.post(
            "/download/excel",
            data={
                "result_json": json.dumps(minimal_result),
                "strategy_name": "sma",
            },
        )
        self.assertEqual(rv.status_code, 200)
        self.assertGreater(len(rv.data), 500)

    def test_download_excel_filename_contains_components(self):
        """Verify /download/excel Content-Disposition filename has correct parts."""
        result = self._make_result()
        rv = self.client.post(
            "/download/excel",
            data={
                "result_json": json.dumps(result),
                "strategy_name": "mean_cost",
            },
        )
        disposition = rv.headers.get("Content-Disposition", "")
        # Filename should contain stock code, strategy name, date range
        self.assertIn("600900", disposition)
        self.assertIn("mean_cost", disposition)
        self.assertIn("2023-01-03", disposition)
        self.assertIn("2023-01-31", disposition)

    def test_download_pdf_filename_contains_components(self):
        """Verify /download/pdf Content-Disposition filename has correct parts."""
        result = self._make_result()
        rv = self.client.post(
            "/download/pdf",
            data={
                "result_json": json.dumps(result),
                "strategy_name": "mean_cost",
            },
        )
        disposition = rv.headers.get("Content-Disposition", "")
        self.assertIn("600900", disposition)
        self.assertIn("mean_cost", disposition)
        self.assertIn(".pdf", disposition)


if __name__ == "__main__":
    unittest.main()
