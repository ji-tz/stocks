"""Unit tests for trader/export.py — Excel export, PDF data prep, filename generation.

Owned by: ITEST (per AGENTS.md)
"""

import json
import tempfile
import unittest
from unittest.mock import patch


class TestExportToExcel(unittest.TestCase):
    """Test export_to_excel() — Excel file generation from backtest results."""

    def _make_backtest_result(self, **overrides) -> dict:
        """Create a minimal backtest result dict for testing."""
        defaults = {
            "symbol": "600900",
            "start_date": "2023-01-01",
            "end_date": "2023-01-31",
            "init_cash": 100000.0,
            "cash": 95000.0,
            "total_value": 105000.0,
            "realized_pl": 3000.0,
            "unrealized_pl": 2000.0,
            "shares": 500,
            "last_price": 22.0,
            "trades": 3,
            "trades_list": [
                {
                    "date": "2023-01-03",
                    "action": "buy",
                    "price": 20.0,
                    "shares": 300,
                    "cash": 94000.0,
                    "shares_after": 300,
                    "realized_pl": 0,
                    "source": "signal",
                },
                {
                    "date": "2023-01-10",
                    "action": "buy",
                    "price": 21.0,
                    "shares": 200,
                    "cash": 73900.0,
                    "shares_after": 500,
                    "realized_pl": 0,
                    "source": "signal",
                },
                {
                    "date": "2023-01-20",
                    "action": "sell",
                    "price": 22.0,
                    "shares": 100,
                    "cash": 75900.0,
                    "shares_after": 400,
                    "realized_pl": 200.0,
                    "source": "signal",
                },
            ],
            "history": [
                {"date": "2023-01-03", "cash": 94000.0, "shares": 300, "avg_cost": 20.0,
                 "last_price": 20.5, "market_value": 6150.0, "total_value": 100150.0},
                {"date": "2023-01-10", "cash": 73900.0, "shares": 500, "avg_cost": 20.4,
                 "last_price": 21.0, "market_value": 10500.0, "total_value": 84400.0},
                {"date": "2023-01-20", "cash": 75900.0, "shares": 400, "avg_cost": 20.5,
                 "last_price": 22.0, "market_value": 8800.0, "total_value": 84700.0},
            ],
        }
        defaults.update(overrides)
        return defaults

    def test_export_excel_creates_file(self):
        """Verify export_to_excel creates a non-empty .xlsx file."""
        from trader.export import export_to_excel

        result = self._make_backtest_result()
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            output_path = export_to_excel(
                backtest_result=result,
                output_path=tmp.name,
                strategy_name="mean_cost",
            )
            self.assertEqual(output_path, tmp.name)

        # Verify file exists and has content
        import os
        self.assertTrue(os.path.exists(output_path))
        self.assertGreater(os.path.getsize(output_path), 0)
        os.unlink(output_path)

    def test_export_excel_has_three_sheets(self):
        """Verify the generated .xlsx has correct sheet names."""
        from trader.export import export_to_excel

        result = self._make_backtest_result()
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            export_to_excel(result, tmp.name, strategy_name="sma")
            import openpyxl
            wb = openpyxl.load_workbook(tmp.name)
            self.assertEqual(wb.sheetnames, ["交易明细", "持仓变化", "汇总统计"])
            wb.close()
        import os
        os.unlink(tmp.name)

    def test_export_excel_trade_details_sheet(self):
        """Verify 交易明细 sheet contains trade records."""
        from trader.export import export_to_excel

        result = self._make_backtest_result()
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            export_to_excel(result, tmp.name)
            import openpyxl
            wb = openpyxl.load_workbook(tmp.name)
            ws = wb["交易明细"]
            # Header row
            self.assertEqual(ws.cell(1, 1).value, "日期")
            self.assertEqual(ws.cell(1, 2).value, "方向")
            self.assertEqual(ws.cell(1, 3).value, "成交价")
            self.assertEqual(ws.cell(1, 4).value, "数量(股)")
            # Data rows (3 trades)
            self.assertEqual(ws.cell(2, 1).value, "2023-01-03")
            self.assertEqual(ws.cell(2, 2).value, "买入")
            self.assertEqual(ws.cell(3, 2).value, "买入")
            self.assertEqual(ws.cell(4, 2).value, "卖出")
            wb.close()
        import os
        os.unlink(tmp.name)

    def test_export_excel_summary_sheet(self):
        """Verify 汇总统计 sheet contains key metrics."""
        from trader.export import export_to_excel

        result = self._make_backtest_result(init_cash=100000.0, realized_pl=3000.0, unrealized_pl=2000.0)
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            export_to_excel(result, tmp.name, strategy_name="mean_cost")
            import openpyxl
            wb = openpyxl.load_workbook(tmp.name)
            ws = wb["汇总统计"]
            # Read key-value pairs
            data = {}
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
                if row[0] and row[1] is not None:
                    data[str(row[0])] = row[1]
            self.assertEqual(data.get("股票代码"), "600900")
            self.assertEqual(data.get("策略名称"), "mean_cost")
            self.assertEqual(data.get("初始资金"), 100000.0)
            self.assertEqual(data.get("交易次数"), 3)
            self.assertEqual(data.get("总盈亏"), 5000.0)  # realized_pl + unrealized_pl
            wb.close()
        import os
        os.unlink(tmp.name)

    def test_export_excel_history_sheet(self):
        """Verify 持仓变化 sheet contains history records."""
        from trader.export import export_to_excel

        result = self._make_backtest_result()
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            export_to_excel(result, tmp.name)
            import openpyxl
            wb = openpyxl.load_workbook(tmp.name)
            ws = wb["持仓变化"]
            headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
            self.assertIn("日期", headers)
            self.assertIn("现金", headers)
            self.assertIn("持仓(股)", headers)
            self.assertIn("总资产", headers)
            # Should have 3 history rows
            self.assertGreaterEqual(ws.max_row, 4)
            wb.close()
        import os
        os.unlink(tmp.name)

    def test_export_excel_empty_trades_list(self):
        """Verify export succeeds with empty trades_list."""
        from trader.export import export_to_excel

        result = self._make_backtest_result(trades_list=[])
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            output_path = export_to_excel(result, tmp.name)
            self.assertTrue(output_path)
            import openpyxl
            wb = openpyxl.load_workbook(tmp.name)
            ws = wb["交易明细"]
            # Only header row, no data
            self.assertEqual(ws.max_row, 1)
            wb.close()
        import os
        os.unlink(tmp.name)

    def test_export_excel_empty_history(self):
        """Verify export succeeds with empty history."""
        from trader.export import export_to_excel

        result = self._make_backtest_result(history=[])
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            output_path = export_to_excel(result, tmp.name)
            self.assertTrue(output_path)  # Should not crash
        import os
        os.unlink(tmp.name)

    def test_export_excel_minimal_result(self):
        """Verify export handles a minimal result dict gracefully."""
        from trader.export import export_to_excel

        minimal = {"symbol": "TEST"}
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            output_path = export_to_excel(minimal, tmp.name)
            self.assertTrue(output_path)
            import openpyxl
            wb = openpyxl.load_workbook(tmp.name)
            # Should still have 3 sheets
            self.assertEqual(wb.sheetnames, ["交易明细", "持仓变化", "汇总统计"])
            wb.close()
        import os
        os.unlink(tmp.name)


class TestPreparePdfData(unittest.TestCase):
    """Test prepare_pdf_data() — structured data extraction for PDF generation."""

    def _make_result(self, **overrides) -> dict:
        defaults = {
            "symbol": "600900",
            "start_date": "2023-01-01",
            "end_date": "2023-12-31",
            "init_cash": 100000.0,
            "cash": 95000.0,
            "total_value": 120000.0,
            "realized_pl": 15000.0,
            "unrealized_pl": 5000.0,
            "shares": 500,
            "last_price": 24.0,
            "trades": 5,
            "trades_list": [
                {"date": "2023-01-10", "action": "buy", "price": 20.0, "shares": 500, "realized_pl": 0},
                {"date": "2023-06-15", "action": "sell", "price": 24.0, "shares": 200, "realized_pl": 2000},
            ],
            "history": [
                {"date": "2023-01-10", "total_value": 100000.0},
                {"date": "2023-03-15", "total_value": 105000.0},
                {"date": "2023-06-15", "total_value": 110000.0},
                {"date": "2023-09-15", "total_value": 115000.0},
                {"date": "2023-12-31", "total_value": 120000.0},
            ],
        }
        defaults.update(overrides)
        return defaults

    def test_prepare_pdf_data_has_summary(self):
        """Verify PDF data includes summary section."""
        from trader.export import prepare_pdf_data

        result = self._make_result()
        pdf_data = prepare_pdf_data(result, "mean_cost")
        self.assertIn("summary", pdf_data)
        self.assertEqual(pdf_data["summary"]["symbol"], "600900")
        self.assertEqual(pdf_data["summary"]["strategy_name"], "mean_cost")
        self.assertEqual(pdf_data["summary"]["init_cash"], 100000.0)

    def test_prepare_pdf_data_has_metrics(self):
        """Verify PDF data includes metrics with calculated values."""
        from trader.export import prepare_pdf_data

        result = self._make_result()
        pdf_data = prepare_pdf_data(result)
        metrics = pdf_data["metrics"]
        self.assertIn("total_return_rate", metrics)
        self.assertIn("annualized_return", metrics)
        self.assertIn("max_drawdown", metrics)
        self.assertIn("sharpe_ratio", metrics)
        self.assertIn("total_pl", metrics)

        # total_pl = realized_pl + unrealized_pl
        self.assertEqual(metrics["total_pl"], 20000.0)

        # total_return_rate = total_pl / init_cash
        self.assertAlmostEqual(metrics["total_return_rate"], 0.2)

    def test_prepare_pdf_data_has_trades(self):
        """Verify PDF data includes trade records."""
        from trader.export import prepare_pdf_data

        result = self._make_result()
        pdf_data = prepare_pdf_data(result)
        self.assertIn("trades", pdf_data)
        self.assertEqual(len(pdf_data["trades"]), 2)

    def test_prepare_pdf_data_empty_trades(self):
        """Verify PDF data handles empty trades gracefully."""
        from trader.export import prepare_pdf_data

        result = self._make_result(trades_list=[])
        pdf_data = prepare_pdf_data(result)
        self.assertEqual(pdf_data["trades"], [])

    def test_prepare_pdf_data_empty_history(self):
        """Verify PDF data with empty history has zero drawdown."""
        from trader.export import prepare_pdf_data

        result = self._make_result(history=[])
        pdf_data = prepare_pdf_data(result)
        self.assertEqual(pdf_data["metrics"]["max_drawdown"], 0.0)
        self.assertEqual(pdf_data["metrics"]["sharpe_ratio"], 0.0)

    def test_prepare_pdf_data_init_cash_zero(self):
        """Verify PDF data handles zero init_cash without division errors."""
        from trader.export import prepare_pdf_data

        result = self._make_result(init_cash=0)
        pdf_data = prepare_pdf_data(result)
        self.assertEqual(pdf_data["metrics"]["total_return_rate"], 0.0)
        self.assertEqual(pdf_data["metrics"]["annualized_return"], 0.0)

    def test_prepare_pdf_data_single_day_history(self):
        """Verify PDF data handles single history entry (no daily returns)."""
        from trader.export import prepare_pdf_data

        result = self._make_result(history=[{"date": "2023-01-10", "total_value": 100000.0}])
        pdf_data = prepare_pdf_data(result)
        # sharpe needs at least 2 history points
        self.assertEqual(pdf_data["metrics"]["sharpe_ratio"], 0.0)

    def test_prepare_pdf_data_max_drawdown_calculation(self):
        """Verify max_drawdown is correctly calculated from history."""
        from trader.export import prepare_pdf_data

        # History with a peak then a drop
        result = self._make_result(history=[
            {"date": "2023-01-01", "total_value": 100000.0},
            {"date": "2023-03-01", "total_value": 120000.0},  # peak
            {"date": "2023-06-01", "total_value": 90000.0},   # 25% drawdown
            {"date": "2023-09-01", "total_value": 110000.0},
        ])
        pdf_data = prepare_pdf_data(result)
        # Max drawdown = (120000 - 90000) / 120000 = 0.25
        self.assertAlmostEqual(pdf_data["metrics"]["max_drawdown"], 0.25, places=4)

    def test_prepare_pdf_data_minimal_result(self):
        """Verify prepare_pdf_data handles a dict with only symbol."""
        from trader.export import prepare_pdf_data

        minimal = {"symbol": "TEST"}
        pdf_data = prepare_pdf_data(minimal)
        self.assertEqual(pdf_data["summary"]["symbol"], "TEST")
        self.assertEqual(pdf_data["summary"]["strategy_name"], "")


class TestGenerateFilename(unittest.TestCase):
    """Test generate_filename() — output file naming."""

    def test_generate_filename_basic(self):
        """Verify basic filename format."""
        from trader.export import generate_filename

        name = generate_filename("600900", "mean_cost", "20230101", "20231231")
        self.assertEqual(name, "600900_mean_cost_20230101_20231231.xlsx")

    def test_generate_filename_pdf_ext(self):
        """Verify PDF extension works."""
        from trader.export import generate_filename

        name = generate_filename("600900", "sma", "20230101", "20231231", ext="pdf")
        self.assertEqual(name, "600900_sma_20230101_20231231.pdf")

    def test_generate_filename_no_dates(self):
        """Verify filename without dates."""
        from trader.export import generate_filename

        name = generate_filename("600900", "mean_cost", "", "")
        self.assertEqual(name, "600900_mean_cost.xlsx")

    def test_generate_filename_no_strategy(self):
        """Verify filename falls back when strategy name is empty."""
        from trader.export import generate_filename

        name = generate_filename("600900", "", "20230101", "20231231")
        self.assertEqual(name, "600900_backtest_20230101_20231231.xlsx")

    def test_generate_filename_slash_in_strategy(self):
        """Verify slashes in strategy name are sanitized."""
        from trader.export import generate_filename

        name = generate_filename("600900", "up/down", "20230101", "20231231")
        self.assertNotIn("/", name)

    def test_generate_filename_no_ext(self):
        """Verify default extension is xlsx."""
        from trader.export import generate_filename

        name = generate_filename("600900", "test", "20230101", "20231231", ext="")
        # Should just have the trailing dot
        self.assertTrue(name.endswith("."))


if __name__ == "__main__":
    unittest.main()
