"""Integration tests for trader/export.py — verifies actual file creation and content.

Owned by: ITEST (per AGENTS.md)
"""

import json
import os
import tempfile
import unittest


class TestExportExcelIntegration(unittest.TestCase):
    """Integration tests for export_to_excel with real file I/O."""

    def _make_backtest_result(self) -> dict:
        """Create a realistic backtest result."""
        return {
            "symbol": "600900",
            "start_date": "2023-01-03",
            "end_date": "2023-01-31",
            "init_cash": 100000.0,
            "cash": 75900.0,
            "total_value": 84700.0,
            "realized_pl": 200.0,
            "unrealized_pl": 8800.0,
            "shares": 400,
            "last_price": 22.0,
            "trades": 5,
            "trades_list": [
                {
                    "date": "2023-01-03",
                    "action": "buy",
                    "price": 20.0,
                    "shares": 300,
                    "cash": 94000.0,
                    "shares_after": 300,
                    "realized_pl": 0,
                    "source": "signal",
                },
                {
                    "date": "2023-01-10",
                    "action": "buy",
                    "price": 21.0,
                    "shares": 200,
                    "cash": 73900.0,
                    "shares_after": 500,
                    "realized_pl": 0,
                    "source": "signal",
                },
                {
                    "date": "2023-01-15",
                    "action": "sell",
                    "price": 22.0,
                    "shares": 100,
                    "cash": 75900.0,
                    "shares_after": 400,
                    "realized_pl": 200.0,
                    "source": "signal",
                },
            ],
            "history": [
                {"date": "2023-01-03", "cash": 94000.0, "shares": 300, "avg_cost": 20.0,
                 "last_price": 20.5, "market_value": 6150.0, "total_value": 100150.0},
                {"date": "2023-01-10", "cash": 73900.0, "shares": 500, "avg_cost": 20.4,
                 "last_price": 21.0, "market_value": 10500.0, "total_value": 84400.0},
                {"date": "2023-01-15", "cash": 75900.0, "shares": 400, "avg_cost": 20.5,
                 "last_price": 22.0, "market_value": 8800.0, "total_value": 84700.0},
            ],
        }

    def test_export_excel_real_file(self):
        """Verify real file is created and can be opened as valid .xlsx."""
        from trader.export import export_to_excel

        result = self._make_backtest_result()
        output_path = os.path.join(tempfile.mkdtemp(), "test_export.xlsx")

        try:
            saved_path = export_to_excel(result, output_path, strategy_name="mean_cost")
            self.assertEqual(saved_path, output_path)
            self.assertTrue(os.path.exists(output_path))
            self.assertGreater(os.path.getsize(output_path), 1000)  # Should be >1KB

            # Validate it's a real .xlsx
            import openpyxl
            wb = openpyxl.load_workbook(output_path)
            self.assertEqual(wb.sheetnames, ["交易明细", "持仓变化", "汇总统计"])
            wb.close()
        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)

    def test_export_excel_sheet_content(self):
        """Verify trade details match expected values in the 交易明细 sheet."""
        from trader.export import export_to_excel

        result = self._make_backtest_result()
        output_path = os.path.join(tempfile.mkdtemp(), "test_content.xlsx")

        try:
            export_to_excel(result, output_path, strategy_name="test")
            import openpyxl
            wb = openpyxl.load_workbook(output_path)

            # Check trade details
            ws = wb["交易明细"]
            trade_rows = []
            for r in range(2, ws.max_row + 1):
                row = {
                    "date": ws.cell(r, 1).value,
                    "direction": ws.cell(r, 2).value,
                    "price": ws.cell(r, 3).value,
                    "shares": ws.cell(r, 4).value,
                    "cash_after": ws.cell(r, 5).value,
                    "shares_after": ws.cell(r, 6).value,
                    "pl": ws.cell(r, 7).value,
                }
                trade_rows.append(row)

            self.assertEqual(len(trade_rows), 3)
            # First trade: buy 300 shares at 20
            self.assertEqual(trade_rows[0]["direction"], "买入")
            self.assertEqual(trade_rows[0]["price"], 20.0)
            self.assertEqual(trade_rows[0]["shares"], 300)

            # Check summary sheet
            ws_sum = wb["汇总统计"]
            summary = {}
            for r in range(1, ws_sum.max_row + 1):
                k = ws_sum.cell(r, 1).value
                v = ws_sum.cell(r, 2).value
                if k:
                    summary[str(k)] = v

            self.assertEqual(summary.get("股票代码"), "600900")
            # total_pl = realized_pl + unrealized_pl = 200 + 8800 = 9000
            self.assertEqual(summary.get("总盈亏"), 9000.0)

            wb.close()
        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)

    def test_export_excel_with_auto_filename(self):
        """Verify filename generation integrated with export."""
        from trader.export import export_to_excel, generate_filename

        result = self._make_backtest_result()
        symbol = result["symbol"]
        strategy_name = "mean_cost"
        start_date = result["start_date"]
        end_date = result["end_date"]

        fname = generate_filename(symbol, strategy_name, start_date, end_date)
        output_dir = tempfile.mkdtemp()
        output_path = os.path.join(output_dir, fname)

        try:
            saved = export_to_excel(result, output_path, strategy_name=strategy_name)
            self.assertEqual(os.path.basename(saved), fname)
            self.assertTrue(fname.startswith("600900_mean_cost"))
            self.assertTrue(fname.endswith(".xlsx"))
        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)

    def test_export_excel_style_structure(self):
        """Verify header rows use correct styling."""
        from trader.export import export_to_excel

        result = self._make_backtest_result()
        output_path = os.path.join(tempfile.mkdtemp(), "test_style.xlsx")

        try:
            export_to_excel(result, output_path)
            import openpyxl
            wb = openpyxl.load_workbook(output_path)

            # Check trade sheet header styling
            ws = wb["交易明细"]
            header_cell = ws.cell(1, 1)
            self.assertTrue(header_cell.font.bold)

            # Check summary sheet key styling
            ws_sum = wb["汇总统计"]
            key_cell = ws_sum.cell(1, 1)
            self.assertTrue(key_cell.font.bold)

            wb.close()
        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)

    def test_export_multiple_calls_same_dir(self):
        """Verify multiple export calls work with different filenames in same dir."""
        from trader.export import export_to_excel, generate_filename

        result = self._make_backtest_result()
        output_dir = tempfile.mkdtemp()

        fname1 = generate_filename("600900", "sma", "20230101", "20231231")
        fname2 = generate_filename("600519", "mean_cost", "20230101", "20231231")
        path1 = os.path.join(output_dir, fname1)
        path2 = os.path.join(output_dir, fname2)

        try:
            p1 = export_to_excel(result, path1, strategy_name="sma")
            p2 = export_to_excel(result, path2, strategy_name="mean_cost")
            self.assertTrue(os.path.exists(p1))
            self.assertTrue(os.path.exists(p2))
            self.assertNotEqual(p1, p2)
        finally:
            for p in [path1, path2]:
                if os.path.exists(p):
                    os.unlink(p)


class TestPdfDataIntegration(unittest.TestCase):
    """Integration: prepare_pdf_data + actual PDF generation via fpdf2."""

    def _make_result(self) -> dict:
        return {
            "symbol": "600900",
            "start_date": "2023-01-01",
            "end_date": "2023-12-31",
            "init_cash": 100000.0,
            "cash": 95000.0,
            "total_value": 120000.0,
            "realized_pl": 15000.0,
            "unrealized_pl": 5000.0,
            "shares": 500,
            "last_price": 24.0,
            "trades": 5,
            "trades_list": [
                {"date": "2023-01-10", "action": "buy", "price": 20.0, "shares": 500, "realized_pl": 0},
                {"date": "2023-06-15", "action": "sell", "price": 24.0, "shares": 200, "realized_pl": 2000},
            ],
            "history": [
                {"date": "2023-01-10", "total_value": 100000.0},
                {"date": "2023-06-15", "total_value": 110000.0},
                {"date": "2023-12-31", "total_value": 120000.0},
            ],
        }

    def test_pdf_data_can_generate_pdf(self):
        """Verify PDF data can be used to generate a real PDF."""
        from trader.export import prepare_pdf_data
        from fpdf import FPDF

        result = self._make_result()
        pdf_data = prepare_pdf_data(result, "mean_cost")

        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Helvetica", size=12)
        summary = pdf_data["summary"]
        pdf.cell(0, 10, f"Symbol: {summary['symbol']}")
        metrics = pdf_data["metrics"]
        pdf.cell(0, 10, f"Total Return: {metrics['total_return_rate'] * 100:.1f}%")
        pdf_bytes = pdf.output()

        self.assertGreater(len(pdf_bytes), 100)

    def test_pdf_data_can_generate_pdf_with_trade_table(self):
        """Verify PDF data includes all trade records for PDF table generation."""
        from trader.export import prepare_pdf_data

        result = self._make_result()
        pdf_data = prepare_pdf_data(result)

        trades = pdf_data["trades"]
        self.assertEqual(len(trades), 2)
        for t in trades:
            self.assertIn("date", t)
            self.assertIn("action", t)
            self.assertIn("price", t)
            self.assertIn("shares", t)
            self.assertIn("pl", t)

    def test_pdf_data_annualized_return_reasonable(self):
        """Verify annualized return is positive and reasonable."""
        from trader.export import prepare_pdf_data

        result = self._make_result()
        pdf_data = prepare_pdf_data(result)
        ar = pdf_data["metrics"]["annualized_return"]
        # With 20% return over 1 year, annualized should be ~20%
        self.assertGreater(ar, 0.15)
        self.assertLess(ar, 0.25)


if __name__ == "__main__":
    unittest.main()
